#!/usr/bin/env python

#
# EAlGIS loader: Australian Census 2016; Data Pack 1
#

import re
import os
import glob
import os.path
import openpyxl
import sqlalchemy
import csv
import json
from datetime import datetime
from collections import OrderedDict

from ealgis_common.loaders import RewrittenCSV, CSVLoader
from ealgis_common.util import alistdir, make_logger
from .shapes import SHAPE_LINKAGE, SHAPE_SCHEMA
from .attrs_repair import repair_census_metadata_first_pass, repair_census_metadata, repair_column_series_census_metadata, repair_series_name

logger = make_logger(__name__)


def parseColumnMetadata(table_number, column_name, metadata):
    """
    Parse Census DataPack column metadata into its component values:
        - Type (Row Name)
        - Kind (Column Name)
    """

    def getColumnLabel(kind):
        # Everything before the pipe is the column label, everything after is the series name
        # e.g. "Owner managers of unincorporated enterprises|MALES"
        # Column = Owner managers of unincorporated enterprises | Series = MALES
        return kind if "|" not in kind else kind.split("|")[0]

    def getSeriesName(kind):
        return None if "|" not in kind else kind.split("|")[1]

    def formatSeriesName(seriesName):
        return seriesName.replace(" - ", " ").replace("-", " ").replace("/ ", " ").replace("/", " ").replace("&", "and")

    def formatColumnLabel(columnLabel):
        """ Format the column label (kind) ready for parsing. """
        return columnLabel.strip().replace(": ", " ").replace(":", " ").replace("-", " ").replace("$", "").replace(":", "").replace("\\", "").replace("&", "and").replace("/ ", " ").replace("/", " ").replace("etc.", "etc").replace(", ", " ").replace("_", " ")

    def formatHumanReadableRowLabel(rowLabel, rowType):
        rowLabel = rowLabel.replace("–", "-").strip()

        # Make currency ranges look nicer
        if "$" in rowType:
            match = re.search(
                r"(?P<rangeStart>[0-9]+)\s(?P<rangeEnd>[0-9]+)", rowLabel)
            if match is not None:
                rangeStart = "{:,}".format(int(match.group("rangeStart")))
                rangeEnd = "{:,}".format(int(match.group("rangeEnd")))
                rowLabel = "${}-${}".format(rangeStart, rangeEnd)

        return rowLabel

    def formatHumanReadableColumnLabel(columnLabel):
        return columnLabel.strip()

    # if table_number != "i15":
    #     return {}
    # if column_name != "i2688":
    #     return {}

    metadata_original = metadata.copy()
    metadata = repair_census_metadata_first_pass(table_number, column_name.lower(), metadata)
    seriesName = repair_series_name(table_number, column_name.lower(), metadata, getSeriesName(metadata["kind"]))
    metadata = repair_census_metadata(table_number, column_name.lower(), metadata, seriesName)

    columnType = metadata["type"].strip()
    columnLabel = formatColumnLabel(getColumnLabel(metadata["kind"]))

    # Special case for B02 and P02 - Selected Medians and Averages
    if table_number == "g02" or table_number == "i04" or table_number == "t02" or table_number == "p02":
        # No named columns here, just rows
        metadata["seriesName"] = None
        metadata["kind"] = ""

    elif seriesName is not None:
        # {SERIES NAME} {ROW LABEL} {COLUMN LABEL}
        # e.g. Persons Speaks other language and speaks English Total Year of arrival 2010
        # Series = Persons, Row = Speaks other language and speaks English Total, Column = Year of arrival 2010
        seriesName = formatSeriesName(seriesName)
        regex = "(?P<seriesName>{seriesName}) (?P<rowLabel>[A-z0-9\s]+) (?P<columnLabel>{columnLabel})".format(seriesName=seriesName.replace(":", ""), columnLabel=columnLabel)
        match = re.search(regex, columnType, re.IGNORECASE)

        if match is not None:
            metadata["seriesName"] = seriesName
            metadata["type"] = formatHumanReadableRowLabel(match.group("rowLabel"), metadata["type"])
            metadata["kind"] = formatHumanReadableColumnLabel(getColumnLabel(metadata["kind"]))
        # else:
        #     print("regex 1", regex)
        #     print("columnType", columnType)

    else:
        # {ROW LABEL} {COLUMN LABEL}
        # e.g. 150 299 Dwelling structure Flat unit or apartment In a 1 or 2 storey block
        # Row = 150 299, Column = Dwelling structure Flat unit or apartment In a 1 or 2 storey block
        regex = "(?P<rowLabel>[A-z0-9\s]+) (?P<columnLabel>{columnLabel})".format(columnLabel=columnLabel)
        match = re.search(regex, columnType, re.IGNORECASE)

        if match is not None:
            metadata["seriesName"] = None
            metadata["type"] = formatHumanReadableRowLabel(match.group("rowLabel"), metadata["type"])
            metadata["kind"] = formatHumanReadableColumnLabel(metadata["kind"])
        # else:
        #     print("regex 2", regex)
        #     print("columnType", columnType)

    if "seriesName" not in metadata:
        raise Exception("[{table_number}] Error parsing column {column_name}: Use \"{type}\" to fix \"{kind}\". Column Label is \"{column_label}\".".format(table_number=table_number.upper(), column_name=column_name, kind=metadata_original["kind"], type=metadata_original["type"].strip().replace("_", " "), column_label=columnLabel))

    # Discard - seriesName is now only used to validate parsing
    del metadata["seriesName"]

    return metadata


def load_metadata_table_serises(loader, census_dir, xlsx_name):
    """
    Parse Census metadata to extract the serises in each table.

    A series represents each set of data within a datapack, e.g.
    Males, Females, Persons

    Returns -
    col_meta[table_number][seriseName] = {
        "columns": [], # The Ids of the columns in a series.
        "datapackNames": [], # The names of the DataPack files (e.g. B12B, B12C) containing the columns for a series.
    }
    """

    def getSeriesName(kind):
        return None if "|" not in kind else kind.split("|")[1]

    col_meta = {}
    col_mapping = {}

    fname = os.path.join(census_dir + '/Metadata/', xlsx_name)
    logger.info("parsing metadata: %s" % (fname))
    wb = openpyxl.load_workbook(fname, read_only=True)

    def sheet_data(sheet):
        return ([t.value for t in r] for r in sheet.iter_rows() if r[0].value is not None)

    def skip_to_descriptors(it):
        for row in it:
            if row[0] != "Cell descriptors":
                next(it)
            else:
                break

    sheet_iter = sheet_data(wb.worksheets[1])
    skip_to_descriptors(sheet_iter)
    for row in sheet_iter:
        name = row[0]
        # Bodge bodge. Fix in skip_to_descriptors()
        # We need to skip twice to get past the header rows
        if not name or name == "Sequential":
            continue
        name = name.lower()
        column_name, short_name, long_name, datapack_file, profile_table, column_heading = row[0:6]

        m = re.match('^([A-Za-z]+[0-9]+)([a-z]+)?$', datapack_file.lower())
        table_number = m.groups()[0]  # b46a -> b46

        column_heading = repair_column_series_census_metadata(table_number, column_name.lower(), str(column_heading).strip())
        seriseName = getSeriesName(column_heading)

        col_mapping[short_name] = column_name

        if seriseName is not None:
            if table_number not in col_meta:
                col_meta[table_number] = {}

            if seriseName not in col_meta[table_number]:
                col_meta[table_number][seriseName] = {
                    "columns": [],
                    "columns_seq": [],
                    "datapackNames": [],
                }

            col_meta[table_number][seriseName]["columns"].append(short_name)  # The shapefile-esque short column name e.g. Tot_P_M
            col_meta[table_number][seriseName]["columns_seq"].append(column_name)  # The sequence_id e.g. G100

            if datapack_file.lower() not in col_meta[table_number][seriseName]["datapackNames"]:
                col_meta[table_number][seriseName]["datapackNames"].append(datapack_file.lower())
    del wb

    return col_meta, col_mapping


def load_metadata(loader, census_dir, xlsx_name, data_tables, columns_by_series):
    table_meta = {}
    col_meta = {}

    fname = os.path.join(census_dir + '/Metadata/', xlsx_name)
    logger.info("parsing metadata: %s" % (fname))
    wb = openpyxl.load_workbook(fname, read_only=True)

    def sheet_data(sheet):
        return ([t.value for t in r] for r in sheet.iter_rows() if r[0].value is not None)

    def skip(it, n):
        for i in range(n):
            next(it)

    def skip_to_descriptors(it):
        for row in sheet_iter:
            if row[0] != "Cell descriptors":
                next(it)
            else:
                break

    def get_metadata_mapping():
        files = {}
        for json_file in glob.glob(os.path.join("./", "census2016", "metadata_mappings", "*_metadata_mapping.json")):
            with open(json_file, "r") as f:
                files = {**files, **json.load(f)["tables"]}
        return files

    def get_topic_to_table_mapping():
        mapping = {}
        for json_file in glob.glob(os.path.join("./", "census2016", "metadata_mappings", "*_topic_mapping.json")):
            with open(json_file, "r") as f:
                for topic_name, tables in json.load(f).items():
                    for table_number in tables:
                        table_number = table_number.upper()
                        if table_number not in mapping:
                            mapping[table_number] = []
                        mapping[table_number].append(topic_name)
        return mapping

    sheet_iter = sheet_data(wb.worksheets[0])
    skip(sheet_iter, 2)
    for row in sheet_iter:
        name = row[0]
        if not name:
            continue
        name = name.lower()
        table_meta[name] = {'type': row[1].strip(), 'kind': row[2].strip() if row[2] is not None else ""}

    sheet_iter = sheet_data(wb.worksheets[1])
    skip_to_descriptors(sheet_iter)
    for row in sheet_iter:
        name = row[0]
        # Bodge bodge. Fix in skip_to_descriptors()
        # We need to skip twice to get past the header rows
        if not name or name == "Sequential":
            continue

        name = name.lower()
        short_name, long_name, datapack_file, profile_table, column_heading = row[1:6]

        datapack_file = datapack_file.lower()
        m = re.match('^([A-Za-z]+[0-9]+)([a-z]+)?$', datapack_file)
        table_number = m.groups()[0]  # b46a -> b46
        if table_number not in col_meta:
            col_meta[table_number] = []

        try:
            col_meta[table_number].append((name, parseColumnMetadata(
                table_number,
                name,
                {'type': str(row[2]).strip(), 'kind': str(row[5]).strip()}
            )))
        except Exception as e:
            if "object has no attribute" in str(e):
                print(name)
                raise e
            logger.error(e)
    del wb

    metadata_mapping = get_metadata_mapping()
    topic_to_table_mapping = get_topic_to_table_mapping()

    for table_name in data_tables:
        datapack_file = table_name.split('_', 1)[0].lower()
        m = re.match('^([A-Za-z]+[0-9]+)(s[0-9]{1,2})?$', datapack_file)
        table_number = m.groups()[0]  # b46a -> b46
        series_id = int(m.groups()[1][1:]) if m.groups()[1] is not None else None  # Just a number that increments from 1
        meta = table_meta[table_number]
        meta["series"] = None
        meta["family"] = table_number

        # Merge JSON formatted metadata from the sequential template XLSs
        if table_number.upper() in metadata_mapping:
            meta = {**meta, **metadata_mapping[table_number.upper()]}

        # Merge JSON formatted topic mappings
        if table_number.upper() in topic_to_table_mapping:
            meta["topics"] = topic_to_table_mapping[table_number.upper()]
        else:
            raise Exception("Couldn't find a topic mapping for table '%s'" % (table_number))

        columns = col_meta[table_number]

        if series_id is not None:
            if table_number not in columns_by_series:
                raise Exception("Expected to find serises for {}".format(table_number))

            series_name = list(columns_by_series[table_number].keys())[series_id - 1]
            meta["series"] = series_name

            # Filter all columns for the table down to just those columns in this series
            columns = [(col_name, col) for col_name, col in col_meta[table_number] if col_name.upper() in columns_by_series[table_number][series_name]["columns_seq"]]

        loader.set_table_metadata(table_name, meta)
        loader.register_columns(table_name, columns)


def load_datapacks(loader, census_dir, tmpdir, packname, abbrev, geo_gid_mapping, columns_by_series, col_mapping):
    def get_csv_files():
        files = []
        for geography in alistdir(d):
            logger.info("%s: Geograpy - %s" % (abbrev, geography))

            g = os.path.join(geography, "*.csv")
            csv_files = glob.glob(g)
            if len(csv_files) == 0:
                g = os.path.join(geography, "AUST", "*.csv")
                csv_files = glob.glob(g)
            if len(csv_files) == 0:
                raise Exception("can't find CSV files for `%s'" % geography)
            files += csv_files
        return files

    def get_csv_files_by_geography_and_table():
        csv_files = get_csv_files()
        by_table = {}

        for i, csv_path in enumerate(csv_files):
            if csv_path.endswith(".tmp.csv"):
                continue

            filename = os.path.basename(csv_path)  # 2016Census_I10B_AUS_IREG.csv
            datapack_file = filename.split('_', 1)[1].lower()  # i10b_aus_ireg.csv
            m = re.match('^([A-Za-z]+[0-9]+)([a-z]+)?_.+$', datapack_file)
            table_number = m.groups()[0]  # i10
            filename_parts = filename.split('_')
            if len(filename_parts) == 4:
                geography_name = filename_parts[3].split(".")[0].lower()  # ireg
            else:
                geography_name = "aust"

            if geography_name not in by_table:
                by_table[geography_name] = {}
            if table_number not in by_table[geography_name]:
                by_table[geography_name][table_number] = []
            by_table[geography_name][table_number].append(csv_path)
        return by_table

    def split_datapack_csv_by_series(columns_by_series, table_name, csv_path):
        csv_files = []

        for key, series_name in enumerate(columns_by_series[table_name]):
            with open(csv_path, "r") as merged_csv_file:
                # Open a new reader for each series as a means of resetting the pointer to the start of the file
                reader = csv.DictReader(merged_csv_file)
                # Need to use this rather than the hard-coded "region_id" fieldname from 2011 due
                # to a change in the structure of the CSV datapacks.
                region_id_column_name = reader.fieldnames[0]  # Per SHAPE_LINKAGE (e.g. AUS_CODE_2016)

                series_csv_path = csv_path.replace(table_name.upper(), "{}S{}".format(table_name.upper(), key + 1))
                if not series_csv_path.endswith(".tmp.csv"):
                    series_csv_path = series_csv_path.replace(".csv", ".tmp.csv")

                # https://stackoverflow.com/a/39923823/7368493
                fieldnames = [region_id_column_name] + columns_by_series[table_name][series_name]["columns"]
                fieldnames_set = set(fieldnames)
                # logger.info("Fieldnames ({}): {}".format(len(fieldnames), fieldnames))

                with open(series_csv_path, "w") as f:
                    writer = csv.DictWriter(f, fieldnames)
                    writer.writeheader()

                    for row in reader:
                        # Use a dictionary comprehension to iterate over the key, value pairs
                        # discarding those pairs whose key is not in the set
                        filtered_row = dict(
                            (k, v) for k, v in row.items() if k in fieldnames_set
                        )
                        writer.writerow(filtered_row)

                logger.info("%s-%s: Created CSV file for series '%s' - %s" % (abbrev, table_name.upper(), series_name, os.path.basename(series_csv_path)))
                csv_files.append(series_csv_path)
        del reader

        return csv_files

    def merge_and_get_csv_files_by_table_and_series():
        csv_files_by_geog_and_table = get_csv_files_by_geography_and_table()
        csv_files = []

        for geography_name, tables in csv_files_by_geog_and_table.items():
            # if geography_name != "lga":
            #     continue

            for table_name, csv_paths in csv_files_by_geog_and_table[geography_name].items():
                # Merge the separate profile table/datapack CSVs into a single new  CSV file based on region_id (first column)
                # if table_name != "g06":
                #     continue

                if len(csv_paths) > 1:
                    dicts = []

                    for i, csv_path in enumerate(csv_paths):
                        with open(csv_path, "r") as f:
                            r = csv.reader(f)
                            if i == 0:
                                dicts.append(OrderedDict((row[0], row[1:]) for row in r))
                            else:
                                dicts.append({row[0]: row[1:] for row in r})

                    result = OrderedDict()
                    for d in tuple(dicts):
                        for key, value in d.items():
                            result.setdefault(key, []).extend(value)

                    profiletable_name = os.path.basename(csv_paths[0]).split('_')[1]
                    merged_csv_path = csv_paths[0].replace("_{}_".format(profiletable_name), "_{}_".format(table_name.upper())).replace(".csv", ".tmp.csv")
                    with open(merged_csv_path, "w") as f:
                        w = csv.writer(f)
                        for key, value in result.items():
                            w.writerow([key] + value)

                    # Some tables are large (and have multiple datapacks), but no serises (e.g. X03)
                    # For these tables we just merge into one combined CSV file...
                    if table_name not in columns_by_series:
                        logger.info("%s: Merged datapack CSV files - %s" % (abbrev, ", ".join([os.path.basename(i) for i in csv_paths])))
                        csv_files.append(merged_csv_path)
                    else:
                        # ...but others are large and DO have serises (e.g. X01)
                        # These we will also merge into one combined CSV file,
                        # then split our merged file into separate CSVs for each
                        # series in the datapack
                        split_csv_files = split_datapack_csv_by_series(columns_by_series, table_name, merged_csv_path)
                        logger.info("%s: Split multiple datapack CSV files - %s" % (abbrev, ", ".join([os.path.basename(i) for i in split_csv_files])))
                        csv_files += split_csv_files

                        # Remove temporary merged CSV path - we have individual CSV files for each series now
                        os.remove(merged_csv_path)

                else:
                    # Some tables are small enough to fit multiple serises in a single datapack CSV file (e.g. P05)
                    # So we need to split these into separate CSVs for each series too
                    if table_name in columns_by_series:
                        split_csv_files = split_datapack_csv_by_series(columns_by_series, table_name, csv_paths[0])
                        logger.info("%s: Split single datapack CSV file - %s" % (abbrev, ", ".join([os.path.basename(i) for i in split_csv_files])))
                        csv_files += split_csv_files
                    else:
                        csv_files.append(csv_paths[0])
        return csv_files

    d = os.path.join(census_dir, packname, "2016 Census %s All Geographies for AUST" % abbrev)
    csv_files = merge_and_get_csv_files_by_table_and_series()

    table_re = re.compile(r'^2016Census_(.*?)(.tmp)?.csv$')
    linkage_pending = []
    data_tables = []

    for i, csv_path in enumerate(csv_files):
        logger.info("%s: [%d/%d] %s" % (abbrev, i + 1, len(csv_files), os.path.basename(csv_path)))
        table_name = table_re.match(os.path.split(csv_path)[-1]).groups()[0].lower()

        data_tables.append(table_name)
        decoded = table_name.split('_')

        if len(decoded) == 3:
            census_division = decoded[2]
        else:
            census_division = None

        gid_match = None

        if census_division is not None:
            def make_match_fn():
                lookup = geo_gid_mapping[census_division]

                def _matcher(line, row):
                    if line == 0:
                        # Rewrite the header
                        # Map from "Tot_P_M" (in the CSV header) to "G100" (in the database)
                        return ['gid', 'region_id'] + [col_mapping[v] for v in row[1:]]
                    else:
                        # Data rows
                        if row[0] in lookup:
                            return [str(lookup[row[0]])] + row
                        else:
                            # Handle missing 9799 LGAs in 2016
                            logger.error("failed gid lookup for '%s' for '%s'" % (row[0], census_division))
                            return [row[0]] + row
                return _matcher
        else:
            # For geometry AUST
            def make_match_fn():
                def _matcher(line, row):
                    if line == 0:
                        # rewrite the header
                        return ['gid', 'region_id'] + [col_mapping[v] for v in row[1:]]
                    else:
                        # rewrite the single row with a gid of 1
                        return ["1"] + row
                return _matcher
        gid_match = make_match_fn()

        # normalise the CSV file by reading it in and writing it out again,
        # Postgres is quite pedantic. we also want to add an additional column to it
        with RewrittenCSV(tmpdir, csv_path, gid_match) as norm:
            instance = CSVLoader(loader.dbschema(), table_name, norm.get(), pkey_column=0)
            table_info = instance.load(loader)
            if table_info is not None and census_division is not None:
                linkage_pending.append((table_name, table_info, census_division))

        # Tidy up after ourselves
        if csv_path.endswith(".tmp.csv"):
            os.remove(csv_path)

    # done as another pass to avoid having to re-run the reflection of the entire
    # database for every CSV file loaded (can be thousands)
    with loader.access_schema(SHAPE_SCHEMA) as geo_access:
        for attr_table, table_info, census_division in linkage_pending:
            geo_column, _, _ = SHAPE_LINKAGE[census_division]
            loader.add_geolinkage(
                geo_access,
                census_division, "gid",
                attr_table, "gid")

    return data_tables


def build_geo_gid_mapping(factory):
    shape_access = factory.make_schema_access(SHAPE_SCHEMA)
    geo_gid_mapping = {}
    for census_division in SHAPE_LINKAGE:
        geo_column, geo_cast_required, _ = SHAPE_LINKAGE[census_division]
        geo_cls = shape_access.get_table_class(census_division)
        geo_attr = getattr(geo_cls, geo_column)
        if geo_cast_required is not None:
            inner_col = sqlalchemy.cast(geo_attr, geo_cast_required)
        else:
            inner_col = geo_attr
        lookup = {}
        for gid, match in shape_access.session.query(geo_cls.gid, inner_col).all():
            lookup[str(match)] = gid
        geo_gid_mapping[census_division] = lookup
    return geo_gid_mapping


def load_attrs(factory, census_dir, tmpdir):
    packages = [
        ("Aboriginal and Torres Strait Islander Peoples Profile", "ATSIP", "Metadata_2016_ATSIP_DataPack.xlsx", "http://www.abs.gov.au/ausstats/abs@.nsf/mf/2069.0.30.002"),
        ("General Community Profile", "GCP", "Metadata_2016_GCP_DataPack.xlsx", "http://www.abs.gov.au/ausstats/abs@.nsf/mf/2001.0"),
        ("Place of Enumeration Profile", "PEP", "Metadata_2016_PEP_DataPack.xlsx", "http://www.abs.gov.au/ausstats%5Cabs@.nsf/0/07ACB32CACD7F2B5CA2573600019C5F3?Opendocument"),
        ("Time Series Profile", "TSP", "Metadata_2016_TSP_DataPack.xlsx", "http://www.abs.gov.au/ausstats/abs@.nsf/mf/2003.0?OpenDocument"),
        ("Working Population Profile", "WPP", "Metadata_2016_WPP_DataPack.xlsx", "http://www.abs.gov.au/ausstats/abs@.nsf/mf/2006.0"),
    ]

    # For dev purposes only - drop schemas on load
    # from sqlalchemy.orm import sessionmaker
    # Session = sessionmaker()
    # Session.configure(bind=factory.engine)
    # session = Session()

    attr_results = []
    geo_gid_mapping = build_geo_gid_mapping(factory)

    for package_name, abbrev, metadata_filename, package_description in packages:
        # session.execute("DROP SCHEMA IF EXISTS aus_census_2016_{} CASCADE".format(abbrev.lower()))
        # session.commit()

        dirname = '2016 ' + package_name
        package_dir = os.path.join(census_dir, dirname)
        schema_name = 'aus_census_2016_' + abbrev.lower()
        loader = factory.make_loader(schema_name)
        loader.add_dependency(SHAPE_SCHEMA)
        loader.set_metadata(
            name=package_name,
            family="ABS Census 2016",
            description=package_description,
            date_published=datetime(2016, 6, 27, 3, 0, 0)  # Set in UTC
        )

        columns_by_series, col_mapping = load_metadata_table_serises(loader, package_dir, metadata_filename)
        data_tables = load_datapacks(loader, census_dir, tmpdir, dirname, abbrev, geo_gid_mapping, columns_by_series, col_mapping)
        load_metadata(loader, package_dir, metadata_filename, data_tables, columns_by_series)
        attr_results.append(loader.result())
    return attr_results
